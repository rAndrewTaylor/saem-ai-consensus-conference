"""Load real WG questions into the platform, replacing demo data.

Usage:
    cd "AI consensus conference"
    python questions/load_questions.py

Reads all files from questions/incoming/, extracts research questions,
and POSTs them to the platform API via the admin endpoint.
"""

import json
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET

# --- Configuration ---
PLATFORM_URL = os.environ.get(
    "PLATFORM_URL",
    "https://saem-ai-consensus-conference-production.up.railway.app"
)
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
INCOMING_DIR = os.path.join(os.path.dirname(__file__), "incoming")

# --- Helpers ---

def extract_docx_text(path):
    """Extract all text from a .docx file."""
    z = zipfile.ZipFile(path)
    doc = z.read("word/document.xml")
    root = ET.fromstring(doc)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    lines = []
    for p in root.findall(".//w:p", ns):
        t = "".join(r.text or "" for r in p.findall(".//w:t", ns))
        if t.strip():
            lines.append(t.strip())
    return lines


def extract_xlsx_questions(path):
    """Extract questions from an Excel file — looks for a column with 'question' in header."""
    try:
        import openpyxl
    except ImportError:
        print("  [!] openpyxl not installed — skipping Excel file")
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    questions = []
    for ws in wb.worksheets:
        headers = [str(c.value or "").lower() for c in ws[1]]
        q_col = None
        for i, h in enumerate(headers):
            if "research question" in h or "question" in h:
                q_col = i
                break
        if q_col is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            if q_col < len(vals) and vals[q_col]:
                text = str(vals[q_col]).strip()
                if len(text) > 20 and "?" in text:
                    questions.append(text)
    return questions


def extract_txt_questions(path):
    """Extract questions from a plain text file (one per line or RQ-prefixed)."""
    with open(path, "r") as f:
        lines = f.readlines()
    questions = []
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        # Strip leading RQ numbering
        cleaned = re.sub(r"^RQ\s*\d+[\.\)]\s*", "", line)
        if len(cleaned) > 20 and "?" in cleaned:
            questions.append(cleaned)
    return questions


def parse_wg1(lines):
    """WG1 has a single blob with numbered questions."""
    text = " ".join(lines)
    # Split on "N. " pattern
    parts = re.split(r"\d+\.\s+", text)
    questions = []
    for p in parts:
        p = p.strip()
        if len(p) > 20:
            # Clean up trailing periods, ensure ends with ?
            p = re.sub(r"[,\.]+$", "", p).strip()
            if not p.endswith("?"):
                p += "?"
            # Split compound questions on ". " if they contain multiple sentences
            # Keep as single question for Delphi
            questions.append(p)
    return questions


def parse_wg3(lines):
    """WG3 has clean RQ1-RQ20 format."""
    questions = []
    for line in lines:
        m = re.match(r"^RQ\d+[\.\s]+(.+)", line)
        if m:
            questions.append(m.group(1).strip())
    return questions


def parse_wg4(lines):
    """WG4 has domain headers then question lines (all end with ?)."""
    questions = []
    skip_headers = {"SAEM 2026", "Working Group 4", "Domain A", "Domain B", "Domain C", "Domain D"}
    for line in lines:
        if any(line.startswith(h) for h in skip_headers):
            continue
        if len(line) > 30 and "?" in line:
            questions.append(line)
    return questions


def parse_wg5(path):
    """WG5 is an Excel file with structured columns."""
    return extract_xlsx_questions(path)


def detect_wg(filename):
    """Guess which WG a file belongs to from the filename."""
    fn = filename.upper()
    if "WG1" in fn:
        return 1
    if "WG2" in fn:
        return 2
    if "WG3" in fn:
        return 3
    if "WG4" in fn:
        return 4
    if "WG5" in fn:
        return 5
    return None


def extract_questions(filepath):
    """Extract questions from any supported file format."""
    fn = os.path.basename(filepath)
    wg = detect_wg(fn)
    ext = os.path.splitext(fn)[1].lower()

    if ext == ".xlsx":
        raw = parse_wg5(filepath) if wg == 5 else extract_xlsx_questions(filepath)
    elif ext == ".docx":
        lines = extract_docx_text(filepath)
        if wg == 1:
            raw = parse_wg1(lines)
        elif wg == 3:
            raw = parse_wg3(lines)
        elif wg == 4:
            raw = parse_wg4(lines)
        else:
            # Generic: grab anything that looks like a question
            raw = [l for l in lines if len(l) > 30 and "?" in l]
    elif ext == ".txt":
        raw = extract_txt_questions(filepath)
    else:
        print(f"  [!] Unsupported format: {ext}")
        return wg, []

    return wg, raw


# --- Main ---

def main():
    if not os.path.isdir(INCOMING_DIR):
        print(f"No incoming directory at {INCOMING_DIR}")
        sys.exit(1)

    files = sorted(os.listdir(INCOMING_DIR))
    if not files:
        print("No files in incoming/")
        sys.exit(1)

    all_questions = {}  # wg_number -> [question_text, ...]

    print(f"Processing {len(files)} files from {INCOMING_DIR}\n")

    for fn in files:
        if fn.startswith("."):
            continue
        path = os.path.join(INCOMING_DIR, fn)
        wg, questions = extract_questions(path)
        if wg is None:
            print(f"  [!] Could not detect WG for {fn} — skipping")
            continue
        print(f"  WG{wg}: {fn} → {len(questions)} questions")
        all_questions.setdefault(wg, []).extend(questions)

    print()
    for wg in sorted(all_questions):
        print(f"WG{wg}: {len(all_questions[wg])} total questions")

    # Write extracted questions to a JSON file for review
    output_path = os.path.join(os.path.dirname(__file__), "extracted_questions.json")
    with open(output_path, "w") as f:
        json.dump(
            {f"WG{k}": v for k, v in sorted(all_questions.items())},
            f,
            indent=2,
        )
    print(f"\nExtracted questions written to {output_path}")
    print("Review the file, then run: python questions/load_questions.py --upload")

    if "--upload" in sys.argv:
        upload_to_platform(all_questions)


def upload_to_platform(all_questions):
    """Upload questions to the live platform via admin API."""
    import urllib.request
    import urllib.error

    if not ADMIN_PASSWORD:
        print("\n[!] Set ADMIN_PASSWORD env var to upload. Example:")
        print('    ADMIN_PASSWORD=yourpass python questions/load_questions.py --upload')
        sys.exit(1)

    # 1. Login
    print(f"\nConnecting to {PLATFORM_URL}...")
    login_data = json.dumps({"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}).encode()
    req = urllib.request.Request(
        f"{PLATFORM_URL}/api/admin/login",
        data=login_data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        resp = urllib.request.urlopen(req)
        token = json.loads(resp.read())["access_token"]
        print("  Authenticated ✓")
    except urllib.error.HTTPError as e:
        print(f"  [!] Login failed: {e.code} {e.read().decode()}")
        sys.exit(1)

    auth_headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    # 2. Clear demo questions (source='demo') by calling demo reset
    print("  Clearing demo data...")
    req = urllib.request.Request(
        f"{PLATFORM_URL}/api/admin/demo/reset",
        data=b"",
        headers=auth_headers,
        method="POST",
    )
    try:
        resp = urllib.request.urlopen(req)
        result = json.loads(resp.read())
        print(f"  Demo data cleared: {result.get('deleted', result)}")
    except urllib.error.HTTPError as e:
        print(f"  [!] Demo reset failed (continuing): {e.read().decode()}")

    # 3. Upload questions per WG
    for wg_number in sorted(all_questions):
        questions = all_questions[wg_number]
        print(f"\n  Uploading WG{wg_number}: {len(questions)} questions...")

        payload = json.dumps([{"text": q, "source": "co_lead"} for q in questions]).encode()
        req = urllib.request.Request(
            f"{PLATFORM_URL}/api/surveys/questions/{wg_number}/bulk",
            data=payload,
            headers=auth_headers,
            method="POST",
        )
        try:
            resp = urllib.request.urlopen(req)
            result = json.loads(resp.read())
            print(f"    ✓ Created {result.get('count', len(questions))} questions")
        except urllib.error.HTTPError as e:
            print(f"    [!] Failed: {e.code} {e.read().decode()}")

    # 4. Activate all questions for Round 1
    print("\n  Activating questions for Round 1...")
    for wg_number in sorted(all_questions):
        req = urllib.request.Request(
            f"{PLATFORM_URL}/api/surveys/questions/{wg_number}/activate",
            data=b"",
            headers=auth_headers,
            method="POST",
        )
        try:
            resp = urllib.request.urlopen(req)
            result = json.loads(resp.read())
            print(f"    WG{wg_number}: {result.get('activated', '?')} questions activated")
        except urllib.error.HTTPError as e:
            print(f"    [!] WG{wg_number} activate failed: {e.read().decode()}")

    print("\n✓ Done. Questions are live for Round 1.")
    print(f"  Platform: {PLATFORM_URL}")


if __name__ == "__main__":
    main()
